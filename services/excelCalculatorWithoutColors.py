from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Side
from models.person import Person
from datetime import date
from typing import List
from calendar import monthrange, month_name

class ExcelService:
    def __init__(self, persons: List[Person], year: int, path: str = "test.xlsx"):
        self.persons = persons
        self.comparing_year = date(year, 1, 1)
        self.path = path
        self.wb = Workbook()
        self.ws = self.wb.active
        
        self.first_column = 2
        self.first_row = 2
        
        
    def create_excel_comparison_table(self):
        self.set_shiftpattern_iterators()
        self.add_content_to_worksheet()
        self.wb.save(self.path)
        
        
    def set_shiftpattern_iterators(self):
        for person in self.persons:
            timespan = abs(person.shiftpattern_start_date - self.comparing_year).days
            if person.shiftpattern_start_date < self.comparing_year:
                person.shiftpattern_iterator = timespan % len(person.shiftsystem.shiftpattern)
            elif person.shiftpattern_start_date > self.comparing_year:
                person.shiftpattern_iterator = len(person.shiftsystem.shiftpattern) - (timespan % len(person.shiftsystem.shiftpattern))
                if person.shiftpattern_iterator == len(person.shiftsystem.shiftpattern):
                    person.shiftpattern_iterator = 0
         
                    
    def add_content_to_worksheet(self):
        first_column = self.first_column
        for month in range(1, 13):
            last_column = first_column + len(self.persons) + 1 # because of DAY | WEEKDAY | PERSON1 | PERSON2 etc.
            days_in_month = monthrange(self.comparing_year.year, month)[1]
            
            self.add_month_name_header(first_column, last_column, month)
            self.add_person_column_header(first_column + 2)
            
            for day in range(1, days_in_month + 1):
                self.add_day_of_month_column(first_column, month, day)
                self.add_weekday_column(first_column + 1, month, day)
                self.add_person_columns(first_column + 2, month, day)
            
            first_column = last_column + 1
            
            self.adjust_column_dimensions()
        self.ws.cell(self.first_row + 1, self.first_column).border = Border(left=Sides.medium)
       
            
    def add_month_name_header(self, first_column, last_column, month):
        self.ws.merge_cells(start_row       = self.first_row, 
                            start_column    = first_column, 
                            end_row         = self.first_row, 
                            end_column      = last_column)
        self.ws.cell(self.first_row, 
                     first_column, 
                     month_name[month])
        self.month_name_header_styling(first_column, last_column)
        
    def month_name_header_styling(self, first_column, last_column):
        self.ws.cell(self.first_row, first_column).border = Border(left=Sides.medium, top=Sides.medium, bottom=Sides.medium) # first cell
        self.ws.cell(self.first_row, last_column).border = Border(top=Sides.medium, right=Sides.medium, bottom=Sides.medium) # last cell
        for column in range(first_column + 1, last_column):
            self.ws.cell(self.first_row, column).border = Border(top=Sides.medium, bottom=Sides.medium)
        
        
    def add_person_column_header(self, current_column):
        for person in self.persons:
            current_cell = self.ws.cell(self.first_row + 1,
                         current_column + self.persons.index(person),
                         person.alias)
            self.person_column_header_styling(person, current_cell)
                
    def person_column_header_styling(self, person, current_cell: Cell):
        current_cell.alignment = Alignment(horizontal="center")
        if self.persons.index(person) == 0 and len(self.persons) == 1: # if there is only one person
             current_cell.border = Border(left=Sides.thin, top=Sides.thin, right=Sides.thin, bottom=Sides.thin)
        elif self.persons.index(person) == 0: # first person
            current_cell.border = Border(left=Sides.thin, top=Sides.thin, bottom=Sides.thin)
        elif self.persons.index(person) == len(self.persons) - 1: # last person
            current_cell.border = Border(top=Sides.thin, right=Sides.medium, bottom=Sides.thin)
        else: # persons between
            current_cell.border = Border(top=Sides.thin, bottom=Sides.thin)
            
            
    def add_day_of_month_column(self, first_column, month, day):
        current_cell = self.ws.cell(self.first_row + 1 + day,
                     first_column,
                     day)
        self.day_of_month_column_styling(current_cell, month, day)
        
    def day_of_month_column_styling(self, current_cell: Cell, month, day):
        current_cell.alignment = Alignment(horizontal="center")
        if day == 1: # first row of the month
            current_cell.border = Border(left=Sides.medium, top=Sides.thin)
        elif monthrange(self.comparing_year.year, month)[1] == day: # last row of the month
            current_cell.border = Border(left=Sides.medium, bottom=Sides.medium)
        else: # other rows
            current_cell.border = Border(left=Sides.medium)
        
        
    def add_weekday_column(self, second_column, month, day):
        current_cell = self.ws.cell(self.first_row + 1 + day,
                     second_column,
                     date(self.comparing_year.year, month, day).strftime("%A")[:3])
        self.weekday_column_styling(current_cell, month, day)
        
    def weekday_column_styling(self, current_cell: Cell, month: int, day: int):
        current_cell.alignment = Alignment(horizontal="center")
        if day == 1: # first row of the month
            current_cell.border = Border( top=Sides.thin, right=Sides.thin)
        elif monthrange(self.comparing_year.year, month)[1] == day: # last row of the month
            current_cell.border = Border(bottom=Sides.medium, right=Sides.thin)
        else: # other rows
            current_cell.border = Border(right=Sides.thin)
    
    
    def add_person_columns(self, third_column, month, day):
        counter = 0
        for person in self.persons:
            current_column = third_column + counter
            current_cell = self.ws.cell(self.first_row + 1 + day,
                                        current_column,
                                        person.shiftsystem.shiftpattern[person.shiftpattern_iterator])
            self.person_column_styling(current_cell, month, day, person)
            person.shiftpattern_iterator += 1
            
            if person.shiftpattern_iterator == len(person.shiftsystem.shiftpattern):
                person.shiftpattern_iterator = 0
            counter += 1
            
    def person_column_styling(self,current_cell: Cell, month: int, day: int, person: Person):
        current_cell.alignment = Alignment(horizontal="center")
        if self.persons.index(person) == len(self.persons) - 1: # last person
            if monthrange(self.comparing_year.year, month)[1] == day: # last monthday
                current_cell.border = Border(right=Sides.medium, bottom=Sides.medium)
            else:
                current_cell.border = Border(right=Sides.medium)
        elif monthrange(self.comparing_year.year, month)[1] == day: # every other person on last monthday
            current_cell.border = Border(bottom=Sides.medium)
        
              
    def adjust_column_dimensions(self):
        columns_per_month = len(self.persons) + 2
        all_columns_of_table = columns_per_month * 12
        for x in range(self.first_column, self.first_column + all_columns_of_table):
            if (x - self.first_column) % columns_per_month < 2:
                self.ws.column_dimensions[self.get_column(x)].width = 5
            else:
                self.ws.column_dimensions[self.get_column(x)].width = 6
        
        
    def get_column(self, num):
        alphabet = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
        if num <= 26:
            return alphabet[num - 1]
        else:
            index = (num - 1) // 26 - 1
            first_letter = alphabet[index]
            second_letter = alphabet[(num % 26) - 1]
            return f"{first_letter}{second_letter}"
        
class Sides:
    black = "ff000000"
    aquamarine = "ffa3e9e2"
    powder_blue = "ff90c6e8"
    
    medium = Side(border_style="medium", color=black)
    thin = Side(border_style="thin", color=black)